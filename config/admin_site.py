from __future__ import annotations

from datetime import date, datetime, time, timedelta

from django.contrib import admin
from django.core.cache import cache
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect
from django.urls import path
from django.utils import timezone

from config.admin_views import DASHBOARD_CACHE_KEY, DASHBOARD_CACHE_TTL, _build_dashboard_context


class ProMasterAdminSite(admin.AdminSite):
    site_header = "МаБибип — Админ-панель"
    site_title = "МаБибип — Админ"
    index_title = "Дашборд"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("dashboard/", self.admin_view(self.dashboard_legacy_redirect), name="dashboard_legacy"),
            path("dashboard/clear-cache/", self.admin_view(self.clear_dashboard_cache), name="dashboard_clear_cache"),
            path("reports/funnel/", self.admin_view(self.reports_funnel), name="reports_funnel"),
            path(
                "reports/funnel.xlsx",
                self.admin_view(self.reports_funnel_xlsx),
                name="reports_funnel_xlsx",
            ),
        ]
        return custom_urls + urls

    def index(self, request: HttpRequest, extra_context=None) -> HttpResponse:  # type: ignore[override]
        if request.GET.get("nocache") == "1":
            ctx = _build_dashboard_context()
        else:
            ctx = cache.get(DASHBOARD_CACHE_KEY)
            if ctx is None:
                ctx = _build_dashboard_context()
                cache.set(DASHBOARD_CACHE_KEY, ctx, DASHBOARD_CACHE_TTL)

        base_ctx = self.each_context(request)
        return super().index(
            request,
            extra_context={
                **base_ctx,
                **ctx,
                "clear_cache_url": self._reverse_admin_url("dashboard_clear_cache"),
            },
        )

    def clear_dashboard_cache(self, request: HttpRequest) -> HttpResponse:
        cache.delete(DASHBOARD_CACHE_KEY)
        return redirect("admin:index")

    def dashboard_legacy_redirect(self, request: HttpRequest) -> HttpResponse:
        return redirect("admin:index")

    def reports_funnel(self, request: HttpRequest) -> HttpResponse:
        if not request.user.is_superuser:
            return redirect("admin:index")

        start, end = self._parse_report_range(request)
        ctx = self._build_funnel_context(start=start, end=end)
        return self._render_admin_template(
            request,
            "admin/reports/funnel.html",
            {
                **self.each_context(request),
                **ctx,
                "start": start,
                "end": end,
                "xlsx_url": self._reverse_admin_url("reports_funnel_xlsx")
                + f"?start={start.isoformat()}&end={end.isoformat()}",
            },
        )

    def reports_funnel_xlsx(self, request: HttpRequest) -> HttpResponse:
        if not request.user.is_superuser:
            return redirect("admin:index")

        start, end = self._parse_report_range(request)
        ctx = self._build_funnel_context(start=start, end=end)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Funnel"

        ws["A1"] = "МаБибип — Отчёт: Воронка"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"] = f"Период: {start.isoformat()} — {end.isoformat()} (включительно)"

        ws["A4"] = "Шаг"
        ws["B4"] = "Пользователей"
        ws["C4"] = "Конверсия от предыдущего шага"
        for cell in ("A4", "B4", "C4"):
            ws[cell].font = Font(bold=True)

        row = 5
        for item in ctx["steps"]:
            ws[f"A{row}"] = item["label"]
            ws[f"B{row}"] = item["count"]
            ws[f"C{row}"] = item["pct_prev"]
            row += 1

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 28
        for r in range(4, row):
            ws[f"B{r}"].alignment = Alignment(horizontal="right")
            ws[f"C{r}"].alignment = Alignment(horizontal="right")

        from io import BytesIO

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="mabibip_funnel_{start.isoformat()}_{end.isoformat()}.xlsx"'
        )
        return resp

    def _parse_report_range(self, request: HttpRequest) -> tuple[date, date]:
        today = timezone.localdate()
        default_start = today - timedelta(days=30)

        raw_start = (request.GET.get("start") or "").strip()
        raw_end = (request.GET.get("end") or "").strip()

        def _parse(raw: str) -> date | None:
            try:
                return date.fromisoformat(raw)
            except Exception:
                return None

        start = _parse(raw_start) or default_start
        end = _parse(raw_end) or today
        if end < start:
            start, end = end, start
        return start, end

    def _build_funnel_context(self, *, start: date, end: date) -> dict:
        from apps.bookings.constants import BookingStatus
        from apps.bookings.models import Booking
        from apps.reviews.models import ModerationStatus, Review
        from apps.users.models import User
        from django.db.models import Q

        # Cohort: users registered in period.
        cohort = User.objects.filter(date_joined__date__gte=start, date_joined__date__lte=end)
        cohort_ids = list(cohort.values_list("id", flat=True))
        registrations = len(cohort_ids)

        if not cohort_ids:
            steps = [
                {"key": "reg", "label": "Регистрация", "count": 0, "pct_prev": "—"},
                {"key": "book", "label": "Создали запись", "count": 0, "pct_prev": "—"},
                {"key": "conf", "label": "СТО подтвердило", "count": 0, "pct_prev": "—"},
                {"key": "done", "label": "Завершено", "count": 0, "pct_prev": "—"},
                {"key": "rev", "label": "Оставили отзыв", "count": 0, "pct_prev": "—"},
            ]
            return {"steps": steps}

        # Bookings created by cohort in period.
        bookings = Booking.objects.filter(
            client_id__in=cohort_ids,
            created_at__date__gte=start,
            created_at__date__lte=end,
        )
        book_users = set(bookings.values_list("client_id", flat=True).distinct())

        # Confirmation/completion based on history (status transition date).
        Hist = Booking.history.model
        status_in_confirmed_or_later = [
            BookingStatus.CONFIRMED,
            BookingStatus.IN_PROGRESS,
            BookingStatus.COMPLETED,
        ]

        confirmed_booking_ids = set(
            Hist.objects.filter(
                id__in=bookings.values_list("id", flat=True),
                status__in=status_in_confirmed_or_later,
                history_date__date__gte=start,
                history_date__date__lte=end,
            ).values_list("id", flat=True)
        )
        confirmed_users = set(
            Booking.objects.filter(id__in=confirmed_booking_ids).values_list("client_id", flat=True).distinct()
        )

        completed_booking_ids = set(
            Hist.objects.filter(
                id__in=bookings.values_list("id", flat=True),
                status=BookingStatus.COMPLETED,
                history_date__date__gte=start,
                history_date__date__lte=end,
            ).values_list("id", flat=True)
        )
        completed_users = set(
            Booking.objects.filter(id__in=completed_booking_ids).values_list("client_id", flat=True).distinct()
        )

        review_users = set(
            Review.objects.filter(
                booking__in=bookings,
                created_at__date__gte=start,
                created_at__date__lte=end,
                moderation_status__in=[ModerationStatus.OK, ModerationStatus.UNDER_REVIEW],
            )
            .values_list("booking__client_id", flat=True)
            .distinct()
        )

        counts = [
            ("reg", "Регистрация", registrations),
            ("book", "Создали запись", len(book_users)),
            ("conf", "СТО подтвердило", len(confirmed_users)),
            ("done", "Завершено", len(completed_users)),
            ("rev", "Оставили отзыв", len(review_users)),
        ]

        steps = []
        prev = None
        for key, label, count in counts:
            if prev is None:
                pct_prev = "—"
            else:
                pct_prev = f"{round(100.0 * count / prev, 1)}%" if prev else "0.0%"
            steps.append({"key": key, "label": label, "count": count, "pct_prev": pct_prev})
            prev = count

        return {"steps": steps}

    def _render_admin_template(self, request: HttpRequest, template_name: str, context: dict) -> HttpResponse:
        from django.template.response import TemplateResponse

        return TemplateResponse(request, template_name, context)

    def _reverse_admin_url(self, name: str) -> str:
        from django.urls import reverse

        return reverse(f"{self.name}:{name}")

