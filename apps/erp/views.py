from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

import json

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from config.admin_views import _build_dashboard_context
from apps.audit.utils import audit_log
from apps.classifieds.erp_stats import platform_classifieds_stats_context


def _is_erp_admin(u) -> bool:
    return bool(u.is_authenticated and u.is_superuser)


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
@require_POST
def city_signal_ack(request: HttpRequest, signal_id: int) -> HttpResponse:
    from apps.core.models import CityExpansionSignal

    obj = get_object_or_404(CityExpansionSignal, pk=int(signal_id))
    obj.acknowledged = True
    obj.save(update_fields=["acknowledged"])
    audit_log(
        request=request,
        event_type="erp.city_expansion_ack",
        action="ack",
        object_label=f"CityExpansionSignal id={obj.pk}",
        payload={"city_label": obj.city_label, "seen_count": int(obj.seen_count)},
    )
    messages.success(request, f"Город «{obj.city_label}» отмечен как «Отлично».")
    return redirect(request.META.get("HTTP_REFERER") or "erp:dashboard")


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def withdrawals_list(request: HttpRequest) -> HttpResponse:
    from apps.billing.models import WithdrawalRequest

    qs = WithdrawalRequest.objects.select_related("wallet", "wallet__user").order_by("-created_at", "-pk")[:300]
    pending = [x for x in qs if x.status == WithdrawalRequest.Status.PENDING]
    return render(
        request,
        "erp/withdrawals.html",
        {"withdrawals": qs, "pending_count": len(pending)},
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
@require_POST
def withdrawal_approve(request: HttpRequest, withdrawal_id: int) -> HttpResponse:
    from apps.billing.models import WalletLedgerEntry, WithdrawalRequest

    obj = get_object_or_404(WithdrawalRequest.objects.select_related("wallet"), pk=int(withdrawal_id))
    if obj.status != WithdrawalRequest.Status.PENDING:
        messages.info(request, "Заявка уже обработана.")
        return redirect("erp:withdrawals")

    obj.status = WithdrawalRequest.Status.APPROVED
    obj.decided_by = request.user
    obj.decided_at = timezone.now()
    obj.save(update_fields=["status", "decided_by", "decided_at"])

    WalletLedgerEntry.objects.create(
        wallet=obj.wallet,
        kind=WalletLedgerEntry.Kind.WITHDRAWAL_APPROVED,
        direction=WalletLedgerEntry.Direction.DEBIT,
        amount=obj.amount,
        currency=obj.currency,
        external_id=f"withdrawal-approved-{obj.pk}",
        payload={"withdrawal_request_id": obj.pk},
    )

    audit_log(
        request=request,
        event_type="billing.withdrawal.approve",
        action="approve",
        object_label=f"WithdrawalRequest id={obj.pk}",
        payload={"wallet_id": obj.wallet_id, "amount": str(obj.amount), "currency": obj.currency},
    )
    messages.success(request, "Заявка на вывод подтверждена.")
    return redirect("erp:withdrawals")


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
@require_POST
def withdrawal_reject(request: HttpRequest, withdrawal_id: int) -> HttpResponse:
    from apps.billing.models import WithdrawalRequest

    obj = get_object_or_404(WithdrawalRequest, pk=int(withdrawal_id))
    if obj.status != WithdrawalRequest.Status.PENDING:
        messages.info(request, "Заявка уже обработана.")
        return redirect("erp:withdrawals")

    comment = (request.POST.get("admin_comment") or "").strip()
    obj.status = WithdrawalRequest.Status.REJECTED
    obj.admin_comment = comment[:300]
    obj.decided_by = request.user
    obj.decided_at = timezone.now()
    obj.save(update_fields=["status", "admin_comment", "decided_by", "decided_at"])

    audit_log(
        request=request,
        event_type="billing.withdrawal.reject",
        action="reject",
        object_label=f"WithdrawalRequest id={obj.pk}",
        payload={"amount": str(obj.amount), "currency": obj.currency, "admin_comment": obj.admin_comment},
    )
    messages.success(request, "Заявка на вывод отклонена.")
    return redirect("erp:withdrawals")


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
@require_POST
def withdrawal_mark_paid(request: HttpRequest, withdrawal_id: int) -> HttpResponse:
    from apps.billing.models import WithdrawalRequest

    obj = get_object_or_404(WithdrawalRequest, pk=int(withdrawal_id))
    if obj.status != WithdrawalRequest.Status.APPROVED:
        messages.info(request, "Отметить «Выплачено» можно только после подтверждения.")
        return redirect("erp:withdrawals")

    obj.status = WithdrawalRequest.Status.PAID
    obj.save(update_fields=["status"])
    audit_log(
        request=request,
        event_type="billing.withdrawal.paid",
        action="paid",
        object_label=f"WithdrawalRequest id={obj.pk}",
        payload={"amount": str(obj.amount), "currency": obj.currency},
    )
    messages.success(request, "Отмечено как «Выплачено».")
    return redirect("erp:withdrawals")


def _parse_range(request: HttpRequest) -> tuple[date, date]:
    today = timezone.localdate()
    default_start = today - timedelta(days=30)

    raw_start = (request.GET.get("start") or "").strip()
    raw_end = (request.GET.get("end") or "").strip()

    def _parse(raw: str) -> date | None:
        try:
            return date.fromisoformat(raw)
        except Exception:
            return None

    start = _parse(raw_start) or default_start
    end = _parse(raw_end) or today
    if end < start:
        start, end = end, start
    return start, end


def _parse_int(raw: str) -> int | None:
    try:
        v = int((raw or "").strip())
        return v
    except Exception:
        return None


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def dashboard(request: HttpRequest) -> HttpResponse:
    ctx = _build_dashboard_context()
    # --- Billing: monthly revenue + payments timeline (platform subscriptions) ---
    from apps.billing.models import PaymentIntent, PaymentIntentStatus
    from django.db.models import Count, Sum
    from django.db.models.functions import TruncDate

    today = timezone.localdate()
    month_start = today - timedelta(days=30)

    intents = PaymentIntent.objects.all()
    succeeded = intents.filter(status=PaymentIntentStatus.SUCCEEDED)
    failed_like = intents.filter(status__in=[PaymentIntentStatus.FAILED, PaymentIntentStatus.CANCELED])
    revenue_30d = (
        succeeded.filter(created_at__date__gte=month_start)
        .aggregate(v=Sum("amount"))
        .get("v")
        or 0
    )
    payments_succeeded_by_day = dict(
        succeeded.filter(created_at__date__gte=month_start)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(v=Sum("amount"))
        .values_list("day", "v")
    )
    payments_failed_cnt_by_day = dict(
        failed_like.filter(created_at__date__gte=month_start)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(c=Count("id"))
        .values_list("day", "c")
    )
    day_list = [today - timedelta(days=i) for i in range(29, -1, -1)]
    ctx["revenue_30d"] = revenue_30d
    ctx["payments_labels_json"] = json.dumps([d.isoformat() for d in day_list])
    ctx["payments_amounts_json"] = json.dumps(
        [float(payments_succeeded_by_day.get(d, 0) or 0) for d in day_list]
    )
    ctx["payments_failed_cnt_json"] = json.dumps(
        [int(payments_failed_cnt_by_day.get(d, 0) or 0) for d in day_list]
    )
    return render(request, "erp/dashboard.html", ctx)


def _build_funnel(*, start: date, end: date) -> dict:
    from apps.bookings.constants import BookingStatus
    from apps.bookings.models import Booking
    from apps.reviews.models import ModerationStatus, Review
    from apps.users.models import User

    cohort = User.objects.filter(date_joined__date__gte=start, date_joined__date__lte=end)
    cohort_ids = list(cohort.values_list("id", flat=True))
    registrations = len(cohort_ids)

    if not cohort_ids:
        steps = [
            {"label": "Регистрация", "count": 0, "pct_prev": "—"},
            {"label": "Создали запись", "count": 0, "pct_prev": "—"},
            {"label": "СТО подтвердило", "count": 0, "pct_prev": "—"},
            {"label": "Завершено", "count": 0, "pct_prev": "—"},
            {"label": "Оставили отзыв", "count": 0, "pct_prev": "—"},
        ]
        return {"steps": steps}

    bookings = Booking.objects.filter(
        client_id__in=cohort_ids,
        created_at__date__gte=start,
        created_at__date__lte=end,
    )
    booking_ids = list(bookings.values_list("id", flat=True))
    book_users = set(bookings.values_list("client_id", flat=True).distinct())

    Hist = Booking.history.model
    confirmed_ids = set(
        Hist.objects.filter(
            id__in=booking_ids,
            status__in=[BookingStatus.CONFIRMED, BookingStatus.IN_PROGRESS, BookingStatus.COMPLETED],
            history_date__date__gte=start,
            history_date__date__lte=end,
        ).values_list("id", flat=True)
    )
    confirmed_users = set(
        Booking.objects.filter(id__in=confirmed_ids).values_list("client_id", flat=True).distinct()
    )

    completed_ids = set(
        Hist.objects.filter(
            id__in=booking_ids,
            status=BookingStatus.COMPLETED,
            history_date__date__gte=start,
            history_date__date__lte=end,
        ).values_list("id", flat=True)
    )
    completed_users = set(
        Booking.objects.filter(id__in=completed_ids).values_list("client_id", flat=True).distinct()
    )

    review_users = set(
        Review.objects.filter(
            booking_id__in=booking_ids,
            created_at__date__gte=start,
            created_at__date__lte=end,
            moderation_status__in=[ModerationStatus.OK, ModerationStatus.UNDER_REVIEW],
        )
        .values_list("booking__client_id", flat=True)
        .distinct()
    )

    counts = [
        ("Регистрация", registrations),
        ("Создали запись", len(book_users)),
        ("СТО подтвердило", len(confirmed_users)),
        ("Завершено", len(completed_users)),
        ("Оставили отзыв", len(review_users)),
    ]
    steps = []
    prev = None
    for label, count in counts:
        if prev is None:
            pct_prev = "—"
        else:
            pct_prev = f"{round(100.0 * count / prev, 1)}%" if prev else "0.0%"
        steps.append({"label": label, "count": count, "pct_prev": pct_prev})
        prev = count
    return {"steps": steps}


@login_required(login_url="/accounts/login/")
def classifieds_stats_report(request: HttpRequest) -> HttpResponse:
    if not _is_erp_admin(request.user):
        return HttpResponse(status=403)
    return render(request, "erp/reports_classifieds.html", platform_classifieds_stats_context())


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def funnel_report(request: HttpRequest) -> HttpResponse:
    start, end = _parse_range(request)
    ctx = _build_funnel(start=start, end=end)
    return render(
        request,
        "erp/reports_funnel.html",
        {
            **ctx,
            "start": start,
            "end": end,
        },
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def funnel_report_xlsx(request: HttpRequest) -> HttpResponse:
    start, end = _parse_range(request)
    ctx = _build_funnel(start=start, end=end)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Funnel"

    ws["A1"] = "МаБибип — Отчёт: Воронка"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Период: {start.isoformat()} — {end.isoformat()} (включительно)"

    ws["A4"] = "Шаг"
    ws["B4"] = "Пользователей"
    ws["C4"] = "Конверсия от предыдущего шага"
    for cell in ("A4", "B4", "C4"):
        ws[cell].font = Font(bold=True)

    row = 5
    for item in ctx["steps"]:
        ws[f"A{row}"] = item["label"]
        ws[f"B{row}"] = item["count"]
        ws[f"C{row}"] = item["pct_prev"]
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    for r in range(4, row):
        ws[f"B{r}"].alignment = Alignment(horizontal="right")
        ws[f"C{r}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="promasterov_funnel_{start.isoformat()}_{end.isoformat()}.xlsx"'
    )
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def activity_report(request: HttpRequest) -> HttpResponse:
    start, end = _parse_range(request)
    ctx = _build_activity(start=start, end=end)
    return render(
        request,
        "erp/reports_activity.html",
        {
            **ctx,
            "start": start,
            "end": end,
        },
    )


def _build_activity(*, start: date, end: date) -> dict:
    from apps.bookings.constants import BookingStatus
    from apps.reviews.models import ModerationStatus
    from apps.stations.models import ServiceStation
    from django.db.models import Avg, Count, Q

    base = ServiceStation.objects.all()
    in_range = Q(bookings__created_at__date__gte=start, bookings__created_at__date__lte=end)

    rev_ok = Q(reviews__moderation_status__in=[ModerationStatus.OK, ModerationStatus.UNDER_REVIEW])

    qs = (
        base.annotate(
            bookings_cnt=Count("bookings", filter=in_range, distinct=True),
            canceled_cnt=Count(
                "bookings",
                filter=in_range & Q(bookings__status=BookingStatus.CANCELED),
                distinct=True,
            ),
            completed_cnt=Count(
                "bookings",
                filter=in_range & Q(bookings__status=BookingStatus.COMPLETED),
                distinct=True,
            ),
            avg_rating=Avg("reviews__rating", filter=rev_ok),
        )
        .filter(bookings_cnt__gt=0)
    )

    top_by_bookings = list(qs.order_by("-bookings_cnt", "-completed_cnt", "name")[:50])
    top_by_canceled = list(qs.order_by("-canceled_cnt", "-bookings_cnt", "name")[:50])
    top_by_rating = list(
        qs.filter(avg_rating__isnull=False).order_by("-avg_rating", "-bookings_cnt", "name")[:50]
    )

    return {
        "top_by_bookings": top_by_bookings,
        "top_by_canceled": top_by_canceled,
        "top_by_rating": top_by_rating,
    }


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def activity_report_xlsx(request: HttpRequest) -> HttpResponse:
    start, end = _parse_range(request)
    ctx = _build_activity(start=start, end=end)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    wb.remove(wb.active)

    def _sheet(title: str, rows):
        ws = wb.create_sheet(title=title[:31])
        ws["A1"] = "МаБибип — Отчёт: Активность СТО"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"] = f"Период: {start.isoformat()} — {end.isoformat()} (включительно)"
        ws["A4"] = "СТО"
        ws["B4"] = "Тип"
        ws["C4"] = "Записей"
        ws["D4"] = "Отмен"
        ws["E4"] = "Заверш"
        ws["F4"] = "Рейтинг"
        for cell in ("A4", "B4", "C4", "D4", "E4", "F4"):
            ws[cell].font = Font(bold=True)

        r = 5
        for st in rows:
            ws[f"A{r}"] = st.name
            ws[f"B{r}"] = getattr(st, "get_executor_kind_display", lambda: st.executor_kind)()
            ws[f"C{r}"] = int(getattr(st, "bookings_cnt", 0) or 0)
            ws[f"D{r}"] = int(getattr(st, "canceled_cnt", 0) or 0)
            ws[f"E{r}"] = int(getattr(st, "completed_cnt", 0) or 0)
            v = getattr(st, "avg_rating", None)
            ws[f"F{r}"] = float(v) if v is not None else ""
            r += 1

        for col, w in {"A": 40, "B": 16, "C": 10, "D": 10, "E": 10, "F": 10}.items():
            ws.column_dimensions[col].width = w
        for rr in range(4, r):
            for col in ("C", "D", "E", "F"):
                ws[f"{col}{rr}"].alignment = Alignment(horizontal="right")

    _sheet("Топ по записям", ctx["top_by_bookings"])
    _sheet("Топ по отменам", ctx["top_by_canceled"])
    _sheet("Топ по рейтингу", ctx["top_by_rating"])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="promasterov_station_activity_{start.isoformat()}_{end.isoformat()}.xlsx"'
    )
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def billing_report(request: HttpRequest) -> HttpResponse:
    from apps.billing.models import PaymentIntent, PaymentIntentStatus, Subscription, SubscriptionStatus
    from django.db.models import Sum

    start, end = _parse_range(request)
    status = (request.GET.get("status") or "").strip()
    provider = (request.GET.get("provider") or "").strip()
    currency = (request.GET.get("currency") or "").strip()
    sub_status = (request.GET.get("sub_status") or "").strip()

    intents = PaymentIntent.objects.select_related("subscription", "subscription__station").filter(
        created_at__date__gte=start, created_at__date__lte=end
    )
    if status:
        intents = intents.filter(status=status)
    if provider:
        intents = intents.filter(provider=provider)
    if currency:
        intents = intents.filter(currency=currency)

    subs = Subscription.objects.select_related("station").all()
    if sub_status:
        subs = subs.filter(status=sub_status)

    intents = intents.order_by("-created_at")[:500]
    subs = subs.order_by("-updated_at")[:500]

    revenue = (
        PaymentIntent.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
            status=PaymentIntentStatus.SUCCEEDED,
        ).aggregate(v=Sum("amount")).get("v")
        or 0
    )

    return render(
        request,
        "erp/reports_billing.html",
        {
            "start": start,
            "end": end,
            "intents": intents,
            "subs": subs,
            "revenue": revenue,
            "status": status,
            "provider": provider,
            "currency": currency,
            "sub_status": sub_status,
            "status_choices": list(PaymentIntentStatus.choices),
            "sub_status_choices": list(SubscriptionStatus.choices),
        },
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def billing_report_xlsx(request: HttpRequest) -> HttpResponse:
    from apps.billing.models import PaymentIntent, Subscription

    start, end = _parse_range(request)
    status = (request.GET.get("status") or "").strip()
    provider = (request.GET.get("provider") or "").strip()
    currency = (request.GET.get("currency") or "").strip()
    sub_status = (request.GET.get("sub_status") or "").strip()

    intents = PaymentIntent.objects.select_related("subscription", "subscription__station").filter(
        created_at__date__gte=start, created_at__date__lte=end
    )
    if status:
        intents = intents.filter(status=status)
    if provider:
        intents = intents.filter(provider=provider)
    if currency:
        intents = intents.filter(currency=currency)
    intents = intents.order_by("-created_at")[:5000]

    subs = Subscription.objects.select_related("station").all()
    if sub_status:
        subs = subs.filter(status=sub_status)
    subs = subs.order_by("-updated_at")[:5000]

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Payments"
    ws1["A1"] = "МаБибип — Отчёт: Платежи"
    ws1["A1"].font = Font(size=14, bold=True)
    ws1["A2"] = f"Период: {start.isoformat()} — {end.isoformat()} (включительно)"
    headers = ["Дата", "Провайдер", "СТО", "Подписка", "Статус", "Сумма", "Валюта", "Provider payment id"]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for it in intents:
        st = getattr(it.subscription, "station", None)
        ws1.cell(row=r, column=1, value=it.created_at.isoformat(sep=" ", timespec="minutes"))
        ws1.cell(row=r, column=2, value=it.provider)
        ws1.cell(row=r, column=3, value=getattr(st, "name", "") if st else "")
        ws1.cell(row=r, column=4, value=str(it.subscription_id))
        ws1.cell(row=r, column=5, value=it.status)
        ws1.cell(row=r, column=6, value=float(it.amount))
        ws1.cell(row=r, column=7, value=it.currency)
        ws1.cell(row=r, column=8, value=it.provider_payment_id)
        r += 1
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 34
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 10
    ws1.column_dimensions["H"].width = 24
    for rr in range(4, r):
        ws1[f"F{rr}"].alignment = Alignment(horizontal="right")

    ws2 = wb.create_sheet(title="Subscriptions")
    ws2["A1"] = "МаБибип — Отчёт: Подписки"
    ws2["A1"].font = Font(size=14, bold=True)
    ws2["A2"] = (
        f"Период: {start.isoformat()} — {end.isoformat()} "
        "(фильтр платежей; подписки — текущее состояние)"
    )
    headers2 = ["Провайдер", "СТО", "Статус", "Период до", "След. списание", "Ошибок подряд", "Последняя ошибка"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for s in subs:
        ws2.cell(row=r, column=1, value=s.provider)
        ws2.cell(row=r, column=2, value=getattr(s.station, "name", ""))
        ws2.cell(row=r, column=3, value=s.status)
        ws2.cell(row=r, column=4, value=s.current_period_end.isoformat() if s.current_period_end else "")
        ws2.cell(
            row=r,
            column=5,
            value=s.next_charge_at.isoformat(sep=" ", timespec="minutes") if s.next_charge_at else "",
        )
        ws2.cell(row=r, column=6, value=int(s.failed_attempts or 0))
        ws2.cell(
            row=r,
            column=7,
            value=s.last_failure_at.isoformat(sep=" ", timespec="minutes") if s.last_failure_at else "",
        )
        r += 1
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 20
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 20

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="promasterov_billing_{start.isoformat()}_{end.isoformat()}.xlsx"'
    )
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def deals_report(request: HttpRequest) -> HttpResponse:
    """
    ERP → Отчёты → Сделки (объявления): фильтры + KPI + таблица.
    """
    from datetime import timedelta

    from django.conf import settings
    from django.db.models import Count, Sum
    from apps.billing.models import ClassifiedsDeal, WithdrawalRequest, WalletLedgerEntry

    start, end = _parse_range(request)
    status = (request.GET.get("status") or "").strip()
    delivery_kind = (request.GET.get("delivery_kind") or "").strip()
    city = (request.GET.get("city") or "").strip()
    seller_id = _parse_int(request.GET.get("seller_id") or "")
    buyer_id = _parse_int(request.GET.get("buyer_id") or "")
    ad_id = _parse_int(request.GET.get("ad_id") or "")
    provider_payment_id = (request.GET.get("provider_payment_id") or "").strip()

    deals = ClassifiedsDeal.objects.select_related("ad", "buyer", "seller").filter(
        created_at__date__gte=start, created_at__date__lte=end
    )
    if status:
        deals = deals.filter(status=status)
    if delivery_kind:
        deals = deals.filter(delivery_kind=delivery_kind)
    if seller_id:
        deals = deals.filter(seller_id=seller_id)
    if buyer_id:
        deals = deals.filter(buyer_id=buyer_id)
    if ad_id:
        deals = deals.filter(ad_id=ad_id)
    if provider_payment_id:
        deals = deals.filter(provider_payment_id=provider_payment_id)
    if city:
        deals = deals.filter(ad__city_label__iexact=city)

    # --- Problem deals (stuck) ---
    now = timezone.now()
    pending_minutes = int(getattr(settings, "ERP_DEAL_PAYMENT_PENDING_MINUTES", 30))
    waiting_hours = int(getattr(settings, "ERP_DEAL_WAITING_SHIPMENT_HOURS", 48))
    pending_cutoff = now - timedelta(minutes=pending_minutes)
    waiting_cutoff = now - timedelta(hours=waiting_hours)

    stuck_payment_qs = ClassifiedsDeal.objects.select_related("ad").filter(
        status__in=[ClassifiedsDeal.Status.CREATED, ClassifiedsDeal.Status.PAYMENT_PENDING],
        created_at__lte=pending_cutoff,
        paid_at__isnull=True,
    )
    stuck_waiting_qs = ClassifiedsDeal.objects.select_related("ad").filter(
        status=ClassifiedsDeal.Status.WAITING_SHIPMENT,
        paid_at__isnull=False,
        paid_at__lte=waiting_cutoff,
    )
    stuck_overdue_confirm_qs = ClassifiedsDeal.objects.select_related("ad").filter(
        status=ClassifiedsDeal.Status.SHIPPED,
        auto_confirm_at__isnull=False,
        auto_confirm_at__lte=now,
    )

    stuck_payment_count = stuck_payment_qs.count()
    stuck_waiting_count = stuck_waiting_qs.count()
    stuck_overdue_confirm_count = stuck_overdue_confirm_qs.count()

    stuck_payment = stuck_payment_qs.order_by("created_at")[:20]
    stuck_waiting = stuck_waiting_qs.order_by("paid_at")[:20]
    stuck_overdue_confirm = stuck_overdue_confirm_qs.order_by("auto_confirm_at")[:20]

    paid_like = [
        ClassifiedsDeal.Status.WAITING_SHIPMENT,
        ClassifiedsDeal.Status.SHIPPED,
        ClassifiedsDeal.Status.BUYER_CONFIRMED,
        ClassifiedsDeal.Status.RELEASED,
        ClassifiedsDeal.Status.REFUND_PENDING,
        ClassifiedsDeal.Status.REFUNDED,
    ]
    gmv_paid = (
        deals.filter(paid_at__isnull=False, status__in=paid_like)
        .aggregate(v=Sum("amount"))
        .get("v")
        or 0
    )
    refunds_sum = deals.filter(status=ClassifiedsDeal.Status.REFUNDED).aggregate(v=Sum("amount")).get("v") or 0
    refunds_count = deals.filter(status=ClassifiedsDeal.Status.REFUNDED).aggregate(c=Count("id")).get("c") or 0
    canceled_before_pay = (
        deals.filter(status=ClassifiedsDeal.Status.CANCELED, paid_at__isnull=True)
        .aggregate(c=Count("id"))
        .get("c")
        or 0
    )

    holds = (
        WalletLedgerEntry.objects.filter(
            kind=WalletLedgerEntry.Kind.DEAL_HOLD,
            created_at__date__gte=start,
            created_at__date__lte=end,
        )
        .aggregate(v=Sum("amount"))
        .get("v")
        or 0
    )
    releases = (
        WalletLedgerEntry.objects.filter(
            kind=WalletLedgerEntry.Kind.DEAL_RELEASE,
            created_at__date__gte=start,
            created_at__date__lte=end,
        )
        .aggregate(v=Sum("amount"))
        .get("v")
        or 0
    )

    wr = WithdrawalRequest.objects.filter(created_at__date__gte=start, created_at__date__lte=end)
    withdrawals_count = wr.aggregate(c=Count("id")).get("c") or 0
    withdrawals_sum = wr.aggregate(v=Sum("amount")).get("v") or 0
    withdrawals_paid_sum = wr.filter(status=WithdrawalRequest.Status.PAID).aggregate(v=Sum("amount")).get("v") or 0

    deals = deals.order_by("-created_at", "-pk")[:500]

    return render(
        request,
        "erp/reports_deals.html",
        {
            "start": start,
            "end": end,
            "deals": deals,
            "status": status,
            "delivery_kind": delivery_kind,
            "city": city,
            "seller_id": seller_id or "",
            "buyer_id": buyer_id or "",
            "ad_id": ad_id or "",
            "provider_payment_id": provider_payment_id,
            "status_choices": list(ClassifiedsDeal.Status.choices),
            "delivery_kind_choices": list(ClassifiedsDeal.DeliveryKind.choices),
            "gmv_paid": gmv_paid,
            "holds": holds,
            "releases": releases,
            "refunds_sum": refunds_sum,
            "refunds_count": refunds_count,
            "canceled_before_pay": canceled_before_pay,
            "withdrawals_count": withdrawals_count,
            "withdrawals_sum": withdrawals_sum,
            "withdrawals_paid_sum": withdrawals_paid_sum,
            "stuck_payment_pending_minutes": pending_minutes,
            "stuck_waiting_shipment_hours": waiting_hours,
            "stuck_payment_count": stuck_payment_count,
            "stuck_waiting_count": stuck_waiting_count,
            "stuck_overdue_confirm_count": stuck_overdue_confirm_count,
            "stuck_payment": stuck_payment,
            "stuck_waiting": stuck_waiting,
            "stuck_overdue_confirm": stuck_overdue_confirm,
        },
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def deals_report_xlsx(request: HttpRequest) -> HttpResponse:
    from apps.billing.models import ClassifiedsDeal

    start, end = _parse_range(request)
    status = (request.GET.get("status") or "").strip()
    delivery_kind = (request.GET.get("delivery_kind") or "").strip()
    city = (request.GET.get("city") or "").strip()
    seller_id = _parse_int(request.GET.get("seller_id") or "")
    buyer_id = _parse_int(request.GET.get("buyer_id") or "")
    ad_id = _parse_int(request.GET.get("ad_id") or "")
    provider_payment_id = (request.GET.get("provider_payment_id") or "").strip()

    deals = ClassifiedsDeal.objects.select_related("ad", "buyer", "seller").filter(
        created_at__date__gte=start, created_at__date__lte=end
    )
    if status:
        deals = deals.filter(status=status)
    if delivery_kind:
        deals = deals.filter(delivery_kind=delivery_kind)
    if seller_id:
        deals = deals.filter(seller_id=seller_id)
    if buyer_id:
        deals = deals.filter(buyer_id=buyer_id)
    if ad_id:
        deals = deals.filter(ad_id=ad_id)
    if provider_payment_id:
        deals = deals.filter(provider_payment_id=provider_payment_id)
    if city:
        deals = deals.filter(ad__city_label__iexact=city)

    deals = deals.order_by("-created_at", "-pk")[:5000]

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"
    ws["A1"] = "МаБибип — Отчёт: Сделки (объявления)"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Период: {start.isoformat()} — {end.isoformat()} (включительно)"

    headers = [
        "ID",
        "Создано",
        "Статус",
        "Доставка",
        "Объявление",
        "Город",
        "Покупатель (id)",
        "Продавец (id)",
        "Сумма",
        "Оплачено",
        "Provider payment id",
        "Автоподтверждение",
        "Отменено",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)

    r = 5
    for d in deals:
        ws.cell(row=r, column=1, value=int(d.pk))
        ws.cell(row=r, column=2, value=d.created_at.isoformat(sep=" ", timespec="minutes"))
        ws.cell(row=r, column=3, value=d.status)
        ws.cell(row=r, column=4, value=d.delivery_kind)
        ws.cell(row=r, column=5, value=f"{d.ad_id} {getattr(d.ad, 'title', '')}")
        ws.cell(row=r, column=6, value=getattr(d.ad, "city_label", "") or "")
        ws.cell(row=r, column=7, value=int(d.buyer_id))
        ws.cell(row=r, column=8, value=int(d.seller_id))
        ws.cell(row=r, column=9, value=float(d.amount))
        ws.cell(row=r, column=10, value=d.paid_at.isoformat(sep=" ", timespec="minutes") if d.paid_at else "")
        ws.cell(row=r, column=11, value=d.provider_payment_id)
        ws.cell(row=r, column=12, value=d.auto_confirm_at.isoformat(sep=" ", timespec="minutes") if d.auto_confirm_at else "")
        ws.cell(row=r, column=13, value=d.canceled_at.isoformat(sep=" ", timespec="minutes") if d.canceled_at else "")
        r += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="tachki_deals_{start.isoformat()}_{end.isoformat()}.xlsx"'
    )
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def deals_users_report(request: HttpRequest) -> HttpResponse:
    """
    ERP → Отчёты → Сделки → Пользователи (топ): buyer/seller.
    """
    from django.db.models import Count, Q, Sum
    from apps.billing.models import ClassifiedsDeal

    start, end = _parse_range(request)
    role = (request.GET.get("role") or "seller").strip()
    if role not in {"seller", "buyer"}:
        role = "seller"

    base = ClassifiedsDeal.objects.filter(created_at__date__gte=start, created_at__date__lte=end)

    who_field = "seller_id" if role == "seller" else "buyer_id"
    paid_like = [
        ClassifiedsDeal.Status.WAITING_SHIPMENT,
        ClassifiedsDeal.Status.SHIPPED,
        ClassifiedsDeal.Status.BUYER_CONFIRMED,
        ClassifiedsDeal.Status.RELEASED,
        ClassifiedsDeal.Status.REFUND_PENDING,
        ClassifiedsDeal.Status.REFUNDED,
    ]

    rows = (
        base.values(who_field)
        .annotate(
            deals_count=Count("id"),
            gmv_paid=Sum("amount", filter=Q(paid_at__isnull=False, status__in=paid_like)),
            refunds_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            refunds_sum=Sum("amount", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            canceled_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.CANCELED)),
        )
        .order_by("-gmv_paid", "-deals_count")[:100]
    )

    # подтянем базовые данные пользователя (email/phone) одним запросом
    from apps.users.models import User

    ids = [r[who_field] for r in rows if r.get(who_field)]
    users = {u.id: u for u in User.objects.filter(id__in=ids).only("id", "phone", "email")}
    for r in rows:
        u = users.get(r.get(who_field))
        r["user"] = u
        r["gmv_paid"] = r.get("gmv_paid") or 0
        r["refunds_sum"] = r.get("refunds_sum") or 0

    return render(
        request,
        "erp/reports_deals_users.html",
        {"start": start, "end": end, "role": role, "rows": rows},
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def deals_users_report_xlsx(request: HttpRequest) -> HttpResponse:
    from django.db.models import Count, Sum
    from django.db.models import Q
    from apps.billing.models import ClassifiedsDeal
    from apps.users.models import User

    start, end = _parse_range(request)
    role = (request.GET.get("role") or "seller").strip()
    if role not in {"seller", "buyer"}:
        role = "seller"
    who_field = "seller_id" if role == "seller" else "buyer_id"
    paid_like = [
        ClassifiedsDeal.Status.WAITING_SHIPMENT,
        ClassifiedsDeal.Status.SHIPPED,
        ClassifiedsDeal.Status.BUYER_CONFIRMED,
        ClassifiedsDeal.Status.RELEASED,
        ClassifiedsDeal.Status.REFUND_PENDING,
        ClassifiedsDeal.Status.REFUNDED,
    ]

    base = ClassifiedsDeal.objects.filter(created_at__date__gte=start, created_at__date__lte=end)
    rows = (
        base.values(who_field)
        .annotate(
            deals_count=Count("id"),
            gmv_paid=Sum("amount", filter=Q(paid_at__isnull=False, status__in=paid_like)),
            refunds_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            refunds_sum=Sum("amount", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            canceled_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.CANCELED)),
        )
        .order_by("-gmv_paid", "-deals_count")[:5000]
    )

    ids = [r[who_field] for r in rows if r.get(who_field)]
    users = {u.id: u for u in User.objects.filter(id__in=ids).only("id", "phone", "email")}

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws["A1"] = "МаБибип — Отчёт: Топ пользователи по сделкам"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Период: {start.isoformat()} — {end.isoformat()} (включительно)"
    ws["A3"] = f"Роль: {role}"

    headers = ["User ID", "Телефон", "Email", "Сделок", "GMV paid", "Refunds cnt", "Refunds sum", "Canceled cnt"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(bold=True)

    r = 6
    for row in rows:
        uid = row.get(who_field)
        u = users.get(uid)
        ws.cell(row=r, column=1, value=int(uid) if uid else "")
        ws.cell(row=r, column=2, value=getattr(u, "phone", "") if u else "")
        ws.cell(row=r, column=3, value=getattr(u, "email", "") if u else "")
        ws.cell(row=r, column=4, value=int(row.get("deals_count") or 0))
        ws.cell(row=r, column=5, value=float(row.get("gmv_paid") or 0))
        ws.cell(row=r, column=6, value=int(row.get("refunds_count") or 0))
        ws.cell(row=r, column=7, value=float(row.get("refunds_sum") or 0))
        ws.cell(row=r, column=8, value=int(row.get("canceled_count") or 0))
        r += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="tachki_deals_users_{role}_{start.isoformat()}_{end.isoformat()}.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def deals_cities_report(request: HttpRequest) -> HttpResponse:
    """
    ERP → Отчёты → Сделки → Города.
    """
    from django.db.models import Count, Q, Sum
    from apps.billing.models import ClassifiedsDeal

    start, end = _parse_range(request)
    base = ClassifiedsDeal.objects.select_related("ad").filter(created_at__date__gte=start, created_at__date__lte=end)

    paid_like = [
        ClassifiedsDeal.Status.WAITING_SHIPMENT,
        ClassifiedsDeal.Status.SHIPPED,
        ClassifiedsDeal.Status.BUYER_CONFIRMED,
        ClassifiedsDeal.Status.RELEASED,
        ClassifiedsDeal.Status.REFUND_PENDING,
        ClassifiedsDeal.Status.REFUNDED,
    ]

    rows = (
        base.values("ad__city_label")
        .annotate(
            deals_count=Count("id"),
            gmv_paid=Sum("amount", filter=Q(paid_at__isnull=False, status__in=paid_like)),
            refunds_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            refunds_sum=Sum("amount", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            canceled_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.CANCELED)),
        )
        .order_by("-gmv_paid", "-deals_count")[:500]
    )
    for r in rows:
        r["city"] = (r.get("ad__city_label") or "").strip() or "—"
        r["gmv_paid"] = r.get("gmv_paid") or 0
        r["refunds_sum"] = r.get("refunds_sum") or 0

    return render(
        request,
        "erp/reports_deals_cities.html",
        {"start": start, "end": end, "rows": rows},
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def deals_cities_report_xlsx(request: HttpRequest) -> HttpResponse:
    from django.db.models import Count, Q, Sum
    from apps.billing.models import ClassifiedsDeal

    start, end = _parse_range(request)
    base = ClassifiedsDeal.objects.select_related("ad").filter(created_at__date__gte=start, created_at__date__lte=end)
    paid_like = [
        ClassifiedsDeal.Status.WAITING_SHIPMENT,
        ClassifiedsDeal.Status.SHIPPED,
        ClassifiedsDeal.Status.BUYER_CONFIRMED,
        ClassifiedsDeal.Status.RELEASED,
        ClassifiedsDeal.Status.REFUND_PENDING,
        ClassifiedsDeal.Status.REFUNDED,
    ]

    rows = (
        base.values("ad__city_label")
        .annotate(
            deals_count=Count("id"),
            gmv_paid=Sum("amount", filter=Q(paid_at__isnull=False, status__in=paid_like)),
            refunds_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            refunds_sum=Sum("amount", filter=Q(status=ClassifiedsDeal.Status.REFUNDED)),
            canceled_count=Count("id", filter=Q(status=ClassifiedsDeal.Status.CANCELED)),
        )
        .order_by("-gmv_paid", "-deals_count")[:5000]
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Cities"
    ws["A1"] = "МаБибип — Отчёт: Сделки по городам"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Период: {start.isoformat()} — {end.isoformat()} (включительно)"

    headers = ["Город", "Сделок", "GMV paid", "Refunds cnt", "Refunds sum", "Canceled cnt"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)

    r = 5
    for row in rows:
        city = (row.get("ad__city_label") or "").strip() or "—"
        ws.cell(row=r, column=1, value=city)
        ws.cell(row=r, column=2, value=int(row.get("deals_count") or 0))
        ws.cell(row=r, column=3, value=float(row.get("gmv_paid") or 0))
        ws.cell(row=r, column=4, value=int(row.get("refunds_count") or 0))
        ws.cell(row=r, column=5, value=float(row.get("refunds_sum") or 0))
        ws.cell(row=r, column=6, value=int(row.get("canceled_count") or 0))
        r += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="tachki_deals_cities_{start.isoformat()}_{end.isoformat()}.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def users_list(request: HttpRequest) -> HttpResponse:
    from apps.users.models import User
    from django.db.models import Count, Q

    q = (request.GET.get("q") or "").strip()
    role = (request.GET.get("role") or "").strip()
    active = (request.GET.get("active") or "").strip()

    qs = User.objects.all().annotate(bookings_cnt=Count("bookings", distinct=True)).order_by("-date_joined")
    if q:
        qs = qs.filter(Q(phone__icontains=q) | Q(email__icontains=q) | Q(id__icontains=q))
    if role == "client":
        qs = qs.filter(is_sto_owner=False)
    elif role == "owner":
        qs = qs.filter(is_sto_owner=True)
    if active in {"0", "1"}:
        qs = qs.filter(is_active=(active == "1"))

    users = list(qs[:500])
    return render(request, "erp/users_list.html", {"users": users, "q": q, "role": role, "active": active})


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def users_list_xlsx(request: HttpRequest) -> HttpResponse:
    from apps.users.models import User
    from django.db.models import Count, Q
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    q = (request.GET.get("q") or "").strip()
    role = (request.GET.get("role") or "").strip()
    active = (request.GET.get("active") or "").strip()

    qs = User.objects.all().annotate(bookings_cnt=Count("bookings", distinct=True)).order_by("-date_joined")
    if q:
        qs = qs.filter(Q(phone__icontains=q) | Q(email__icontains=q) | Q(id__icontains=q))
    if role == "client":
        qs = qs.filter(is_sto_owner=False)
    elif role == "owner":
        qs = qs.filter(is_sto_owner=True)
    if active in {"0", "1"}:
        qs = qs.filter(is_active=(active == "1"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws["A1"] = "МаБибип — ERP экспорт: Пользователи"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Фильтры: q={q or '—'}, role={role or '—'}, active={active or '—'}"
    headers = ["ID", "Телефон", "Email", "Владелец СТО", "Модерация СТО", "Активен", "Записей", "Регистрация"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for u in qs[:5000]:
        ws.cell(row=r, column=1, value=int(u.id))
        ws.cell(row=r, column=2, value=u.phone)
        ws.cell(row=r, column=3, value=u.email or "")
        ws.cell(row=r, column=4, value="Да" if u.is_sto_owner else "Нет")
        ws.cell(row=r, column=5, value=u.sto_moderation_status)
        ws.cell(row=r, column=6, value="Да" if u.is_active else "Нет")
        ws.cell(row=r, column=7, value=int(getattr(u, "bookings_cnt", 0) or 0))
        ws.cell(row=r, column=8, value=u.date_joined.isoformat(sep=" ", timespec="minutes"))
        r += 1
    for col, w in {"A": 8, "B": 18, "C": 26, "D": 12, "E": 16, "F": 10, "G": 10, "H": 20}.items():
        ws.column_dimensions[col].width = w
    for rr in range(4, r):
        ws[f"A{rr}"].alignment = Alignment(horizontal="right")
        ws[f"G{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="erp_users.xlsx"'
    return resp


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_activate(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.users.models import User

    u = User.objects.get(pk=user_id)
    u.is_active = True
    u.save(update_fields=["is_active"])
    audit_log(
        request=request,
        event_type="erp.user",
        action="activate",
        obj=u,
        object_label=f"User id={u.id}",
        payload={},
    )
    return redirect("erp:user_detail", user_id=u.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_deactivate(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.users.models import User

    u = User.objects.get(pk=user_id)
    u.is_active = False
    u.save(update_fields=["is_active"])
    audit_log(
        request=request,
        event_type="erp.user",
        action="deactivate",
        obj=u,
        object_label=f"User id={u.id}",
        payload={},
    )
    return redirect("erp:user_detail", user_id=u.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_set_role(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.users.models import User

    role = (request.POST.get("role") or "").strip()
    u = User.objects.get(pk=user_id)
    old = bool(u.is_sto_owner)
    if role == "owner":
        u.is_sto_owner = True
    elif role == "client":
        u.is_sto_owner = False
    else:
        return redirect("erp:user_detail", user_id=u.id)
    u.save(update_fields=["is_sto_owner"])
    audit_log(
        request=request,
        event_type="erp.user",
        action="set_role",
        obj=u,
        object_label=f"User id={u.id}",
        payload={"from_owner": old, "to": role},
    )
    return redirect("erp:user_detail", user_id=u.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_sto_approve(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.stations.models import ServiceStation
    from apps.users.models import User

    u = User.objects.get(pk=user_id)
    u.is_sto_owner = True
    u.sto_moderation_status = User.StoModerationStatus.APPROVED
    u.save(update_fields=["is_sto_owner", "sto_moderation_status"])
    ServiceStation.objects.filter(owner=u).update(is_active=True)
    audit_log(
        request=request,
        event_type="erp.user_sto_moderation",
        action="approve",
        obj=u,
        object_label=f"User id={u.id}",
        payload={},
    )
    return redirect("erp:user_detail", user_id=u.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_sto_set_moderation(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.users.models import User

    status = (request.POST.get("status") or "").strip()
    allowed = {s for s, _ in User.StoModerationStatus.choices}
    if status not in allowed:
        return redirect("erp:user_detail", user_id=user_id)
    u = User.objects.get(pk=user_id)
    old = u.sto_moderation_status
    u.sto_moderation_status = status
    u.save(update_fields=["sto_moderation_status"])
    audit_log(
        request=request,
        event_type="erp.user_sto_moderation",
        action="set",
        obj=u,
        object_label=f"User id={u.id}",
        payload={"from": old, "to": status},
    )
    return redirect("erp:user_detail", user_id=u.id)


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def bookings_list(request: HttpRequest) -> HttpResponse:
    from apps.bookings.constants import BookingStatus
    from apps.bookings.models import Booking
    from django.db.models import Q

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    qs = Booking.objects.select_related("client", "station", "slot").order_by("-created_at")
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(Q(contact_phone__icontains=q) | Q(station__name__icontains=q) | Q(client__phone__icontains=q))

    bookings = list(qs[:500])
    return render(
        request,
        "erp/bookings_list.html",
        {"bookings": bookings, "q": q, "status": status, "status_choices": list(BookingStatus.choices)},
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def bookings_list_xlsx(request: HttpRequest) -> HttpResponse:
    from apps.bookings.models import Booking
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from django.db.models import Q

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    qs = Booking.objects.select_related("client", "station", "slot").order_by("-created_at")
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(Q(contact_phone__icontains=q) | Q(station__name__icontains=q) | Q(client__phone__icontains=q))

    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    ws["A1"] = "МаБибип — ERP экспорт: Записи"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Фильтры: q={q or '—'}, status={status or '—'}"
    headers = ["ID", "Клиент", "СТО", "Слот", "Статус", "Телефон", "Создано"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for b in qs[:5000]:
        slot = ""
        if b.slot_id:
            slot = f"{b.slot.date} {b.slot.start_time}-{b.slot.end_time}"
        ws.cell(row=r, column=1, value=int(b.id))
        ws.cell(row=r, column=2, value=getattr(b.client, "phone", str(b.client_id)))
        ws.cell(row=r, column=3, value=b.station.name)
        ws.cell(row=r, column=4, value=slot)
        ws.cell(row=r, column=5, value=b.status)
        ws.cell(row=r, column=6, value=b.contact_phone)
        ws.cell(row=r, column=7, value=b.created_at.isoformat(sep=" ", timespec="minutes"))
        r += 1
    for col, w in {"A": 8, "B": 18, "C": 34, "D": 24, "E": 14, "F": 18, "G": 20}.items():
        ws.column_dimensions[col].width = w
    for rr in range(4, r):
        ws[f"A{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="erp_bookings.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def reviews_list(request: HttpRequest) -> HttpResponse:
    from apps.reviews.models import ComplaintStatus, ModerationStatus, Review, ReviewComplaint
    from django.db.models import Q

    q = (request.GET.get("q") or "").strip()
    mod = (request.GET.get("mod") or "").strip()
    comp = (request.GET.get("comp") or "").strip()

    qs = Review.objects.select_related("booking", "booking__station", "booking__client").order_by("-created_at")
    if mod:
        qs = qs.filter(moderation_status=mod)
    if q:
        qs = qs.filter(
            Q(text__icontains=q)
            | Q(booking__station__name__icontains=q)
            | Q(booking__client__phone__icontains=q)
        )
    reviews = list(qs[:500])

    complaints_qs = ReviewComplaint.objects.select_related("review", "station").order_by("-created_at")
    if comp:
        complaints_qs = complaints_qs.filter(status=comp)
    complaints = list(complaints_qs[:300])

    return render(
        request,
        "erp/reviews_list.html",
        {
            "reviews": reviews,
            "complaints": complaints,
            "q": q,
            "mod": mod,
            "comp": comp,
            "mod_choices": list(ModerationStatus.choices),
            "comp_choices": list(ComplaintStatus.choices),
        },
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def reviews_list_xlsx(request: HttpRequest) -> HttpResponse:
    from apps.reviews.models import Review
    from django.db.models import Q
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    q = (request.GET.get("q") or "").strip()
    mod = (request.GET.get("mod") or "").strip()
    qs = Review.objects.select_related("booking", "booking__station", "booking__client").order_by("-created_at")
    if mod:
        qs = qs.filter(moderation_status=mod)
    if q:
        qs = qs.filter(
            Q(text__icontains=q)
            | Q(booking__station__name__icontains=q)
            | Q(booking__client__phone__icontains=q)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"
    ws["A1"] = "МаБибип — ERP экспорт: Отзывы"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Фильтры: q={q or '—'}, moderation={mod or '—'}"
    headers = ["ID", "СТО", "Клиент", "Рейтинг", "Статус", "Дата", "Текст"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for rev in qs[:5000]:
        ws.cell(row=r, column=1, value=int(rev.id))
        ws.cell(row=r, column=2, value=rev.booking.station.name if rev.booking_id else "")
        ws.cell(row=r, column=3, value=rev.booking.client.phone if rev.booking_id else "")
        ws.cell(row=r, column=4, value=int(rev.rating))
        ws.cell(row=r, column=5, value=rev.moderation_status)
        ws.cell(row=r, column=6, value=rev.created_at.date().isoformat())
        ws.cell(row=r, column=7, value=(rev.text or "")[:2000])
        r += 1
    for col, w in {"A": 8, "B": 34, "C": 18, "D": 8, "E": 16, "F": 12, "G": 80}.items():
        ws.column_dimensions[col].width = w
    for rr in range(4, r):
        ws[f"A{rr}"].alignment = Alignment(horizontal="right")
        ws[f"D{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="erp_reviews.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_detail(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.audit.models import AuditLog
    from apps.bookings.models import Booking
    from apps.reviews.models import Review
    from apps.users.models import User

    u = User.objects.get(pk=user_id)
    bookings = (
        Booking.objects.select_related("station", "slot")
        .filter(client=u)
        .order_by("-created_at")[:200]
    )
    reviews = (
        Review.objects.select_related("booking", "booking__station")
        .filter(booking__client=u)
        .order_by("-created_at")[:200]
    )
    audit = AuditLog.objects.filter(actor=u).order_by("-created_at")[:200]
    audit_obj = AuditLog.objects.filter(object_type="users.User", object_id=u.id).order_by("-created_at")[:200]

    return render(
        request,
        "erp/user_detail.html",
        {
            "u": u,
            "bookings": bookings,
            "reviews": reviews,
            "audit": audit,
            "audit_obj": audit_obj,
        },
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_bookings_xlsx(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.bookings.models import Booking
    from apps.users.models import User
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    u = User.objects.get(pk=user_id)
    qs = Booking.objects.select_related("station", "slot").filter(client=u).order_by("-created_at")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    ws["A1"] = "МаБибип — ERP экспорт: Записи пользователя"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Пользователь: {u.id} {u.phone}"
    headers = ["ID", "СТО", "Слот", "Статус", "Телефон", "Создано"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for b in qs[:10000]:
        slot = ""
        if b.slot_id:
            slot = f"{b.slot.date} {b.slot.start_time}-{b.slot.end_time}"
        ws.cell(row=r, column=1, value=int(b.id))
        ws.cell(row=r, column=2, value=b.station.name)
        ws.cell(row=r, column=3, value=slot)
        ws.cell(row=r, column=4, value=b.status)
        ws.cell(row=r, column=5, value=b.contact_phone)
        ws.cell(row=r, column=6, value=b.created_at.isoformat(sep=" ", timespec="minutes"))
        r += 1
    for col, w in {"A": 8, "B": 34, "C": 24, "D": 14, "E": 18, "F": 20}.items():
        ws.column_dimensions[col].width = w
    for rr in range(4, r):
        ws[f"A{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="erp_user_{u.id}_bookings.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def user_reviews_xlsx(request: HttpRequest, user_id: int) -> HttpResponse:
    from apps.reviews.models import Review, ReviewComplaint
    from apps.users.models import User
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    u = User.objects.get(pk=user_id)
    reviews = (
        Review.objects.select_related("booking", "booking__station")
        .filter(booking__client=u)
        .order_by("-created_at")[:10000]
    )
    complaints = (
        ReviewComplaint.objects.select_related("review", "station")
        .filter(review__booking__client=u)
        .order_by("-created_at")[:10000]
    )

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Reviews"
    ws1["A1"] = "МаБибип — ERP экспорт: Отзывы пользователя"
    ws1["A1"].font = Font(size=14, bold=True)
    ws1["A2"] = f"Пользователь: {u.id} {u.phone}"
    headers = ["ID", "СТО", "Рейтинг", "Статус", "Дата", "Текст", "Причина модерации"]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for rev in reviews:
        ws1.cell(row=r, column=1, value=int(rev.id))
        ws1.cell(row=r, column=2, value=rev.booking.station.name if rev.booking_id else "")
        ws1.cell(row=r, column=3, value=int(rev.rating))
        ws1.cell(row=r, column=4, value=rev.moderation_status)
        ws1.cell(row=r, column=5, value=rev.created_at.date().isoformat())
        ws1.cell(row=r, column=6, value=(rev.text or "")[:2000])
        ws1.cell(row=r, column=7, value=rev.moderation_reason or "")
        r += 1
    for col, w in {"A": 8, "B": 34, "C": 8, "D": 16, "E": 12, "F": 80, "G": 40}.items():
        ws1.column_dimensions[col].width = w
    for rr in range(4, r):
        ws1[f"A{rr}"].alignment = Alignment(horizontal="right")
        ws1[f"C{rr}"].alignment = Alignment(horizontal="right")

    ws2 = wb.create_sheet(title="Complaints")
    ws2["A1"] = "МаБибип — ERP экспорт: Жалобы на отзывы пользователя"
    ws2["A1"].font = Font(size=14, bold=True)
    ws2["A2"] = f"Пользователь: {u.id} {u.phone}"
    headers2 = ["ID", "Review ID", "СТО (кто пожаловался)", "Причина", "Статус", "Создано", "Решено"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for cpl in complaints:
        ws2.cell(row=r, column=1, value=int(cpl.id))
        ws2.cell(row=r, column=2, value=int(cpl.review_id))
        ws2.cell(row=r, column=3, value=cpl.station.name if cpl.station_id else "")
        ws2.cell(row=r, column=4, value=cpl.reason)
        ws2.cell(row=r, column=5, value=cpl.status)
        ws2.cell(row=r, column=6, value=cpl.created_at.isoformat(sep=" ", timespec="minutes"))
        ws2.cell(
            row=r,
            column=7,
            value=cpl.resolved_at.isoformat(sep=" ", timespec="minutes") if cpl.resolved_at else "",
        )
        r += 1
    for col, w in {"A": 8, "B": 10, "C": 34, "D": 60, "E": 14, "F": 20, "G": 20}.items():
        ws2.column_dimensions[col].width = w
    for rr in range(4, r):
        ws2[f"A{rr}"].alignment = Alignment(horizontal="right")
        ws2[f"B{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="erp_user_{u.id}_reviews_complaints.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def booking_detail(request: HttpRequest, booking_id: int) -> HttpResponse:
    from apps.bookings.models import Booking
    from apps.bookings.constants import BookingStatus
    from apps.chat.models import ChatRoom, Message

    b = (
        Booking.objects.select_related("client", "station", "station__owner", "slot")
        .get(pk=booking_id)
    )

    Hist = Booking.history.model
    history = list(
        Hist.objects.filter(id=b.id)
        .order_by("-history_date")[:200]
    )

    room = ChatRoom.objects.filter(booking=b).first()
    messages = []
    if room:
        messages = list(
            Message.objects.filter(room=room)
            .select_related("sender")
            .order_by("-created_at", "-pk")[:50]
        )

    review = getattr(b, "review", None)

    return render(
        request,
        "erp/booking_detail.html",
        {
            "b": b,
            "history": history,
            "room": room,
            "messages": list(reversed(messages)),
            "review": review,
            "status_choices": list(BookingStatus.choices),
        },
    )


def _erp_can_set_booking_status(cur: str, new: str) -> bool:
    from apps.bookings.constants import BookingStatus

    if new == BookingStatus.CONFIRMED:
        return cur == BookingStatus.PENDING
    if new == BookingStatus.IN_PROGRESS:
        return cur == BookingStatus.CONFIRMED
    if new == BookingStatus.COMPLETED:
        return cur == BookingStatus.IN_PROGRESS
    if new == BookingStatus.CANCELED:
        return cur in {BookingStatus.PENDING, BookingStatus.CONFIRMED, BookingStatus.IN_PROGRESS}
    return False


def _notify_booking_status_change(*, booking_id: int, old_status: str, new_status: str) -> None:
    from django.db import transaction

    def _send() -> None:
        from apps.bookings import mail as booking_mail
        from apps.bookings.models import Booking

        try:
            b = Booking.objects.select_related("client", "station", "slot").get(pk=booking_id)
            if new_status == "confirmed" and old_status == "pending":
                booking_mail.mail_client_booking_confirmed(b)
            elif new_status == "completed" and old_status == "in_progress":
                booking_mail.mail_client_booking_completed(b)
            elif new_status == "canceled" and old_status in ("pending", "confirmed", "in_progress"):
                booking_mail.mail_client_booking_canceled_by_sto(b)
        except Exception:
            # best-effort; not blocking ERP action
            return

    transaction.on_commit(_send)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def booking_set_status(request: HttpRequest, booking_id: int) -> HttpResponse:
    from django.db import transaction
    from apps.bookings.constants import BookingStatus
    from apps.bookings.models import Booking

    new_status = (request.POST.get("status") or "").strip()
    if new_status not in {s for s, _ in BookingStatus.choices}:
        messages.error(request, "Неверный статус.")
        return redirect("erp:booking_detail", booking_id=booking_id)

    b = Booking.objects.select_related("station", "client").get(pk=booking_id)
    old = b.status
    if not _erp_can_set_booking_status(old, new_status):
        messages.error(request, "Недопустимый переход статуса для текущей записи.")
        audit_log(
            request=request,
            event_type="erp.booking_status",
            action="denied",
            obj=b,
            object_label=f"Booking id={b.id}",
            payload={"from": old, "to": new_status},
        )
        return redirect("erp:booking_detail", booking_id=b.id)

    with transaction.atomic():
        b.status = new_status
        b.save(update_fields=["status"])
        _notify_booking_status_change(booking_id=b.id, old_status=old, new_status=new_status)

    audit_log(
        request=request,
        event_type="erp.booking_status",
        action="set_status",
        obj=b,
        object_label=f"Booking id={b.id} station={b.station_id}",
        payload={"from": old, "to": new_status},
    )
    messages.success(request, f"Статус изменён: {old} → {new_status}.")
    return redirect("erp:booking_detail", booking_id=b.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def booking_emergency_cancel(request: HttpRequest, booking_id: int) -> HttpResponse:
    from django.db import transaction
    from apps.bookings.constants import BookingStatus
    from apps.bookings.models import Booking

    reason = (request.POST.get("reason") or "").strip()[:500]
    if not reason:
        reason = "Экстренная отмена администратором"

    b = Booking.objects.select_related("station", "client").get(pk=booking_id)
    old = b.status
    if not _erp_can_set_booking_status(old, BookingStatus.CANCELED):
        messages.error(request, "Эту запись нельзя отменить в текущем статусе.")
        audit_log(
            request=request,
            event_type="erp.booking_cancel",
            action="denied",
            obj=b,
            object_label=f"Booking id={b.id}",
            payload={"from": old, "reason": reason},
        )
        return redirect("erp:booking_detail", booking_id=b.id)

    with transaction.atomic():
        b.status = BookingStatus.CANCELED
        b.owner_cancel_reason = reason
        b.save(update_fields=["status", "owner_cancel_reason"])
        _notify_booking_status_change(booking_id=b.id, old_status=old, new_status=BookingStatus.CANCELED)

    audit_log(
        request=request,
        event_type="erp.booking_cancel",
        action="emergency_cancel",
        obj=b,
        object_label=f"Booking id={b.id} station={b.station_id}",
        payload={"from": old, "reason": reason},
    )
    messages.success(request, "Запись отменена.")
    return redirect("erp:booking_detail", booking_id=b.id)


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def stations_list(request: HttpRequest) -> HttpResponse:
    from apps.stations.models import ServiceStation
    from django.db.models import Q

    q = (request.GET.get("q") or "").strip()
    active = (request.GET.get("active") or "").strip()

    qs = ServiceStation.objects.select_related("owner", "district").order_by("name")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(address__icontains=q) | Q(owner__phone__icontains=q))
    if active in {"0", "1"}:
        qs = qs.filter(is_active=(active == "1"))

    stations = list(qs[:500])
    return render(request, "erp/stations_list.html", {"stations": stations, "q": q, "active": active})


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def stations_list_xlsx(request: HttpRequest) -> HttpResponse:
    from apps.stations.models import ServiceStation
    from django.db.models import Q
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    q = (request.GET.get("q") or "").strip()
    active = (request.GET.get("active") or "").strip()
    qs = ServiceStation.objects.select_related("owner", "district").order_by("name")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(address__icontains=q) | Q(owner__phone__icontains=q))
    if active in {"0", "1"}:
        qs = qs.filter(is_active=(active == "1"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Stations"
    ws["A1"] = "МаБибип — ERP экспорт: СТО"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Фильтры: q={q or '—'}, active={active or '—'}"
    headers = ["ID", "Название", "Владелец", "Адрес", "Тип", "Тариф", "Оплачено до", "Активна", "Проверен"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for s in qs[:5000]:
        ws.cell(row=r, column=1, value=int(s.id))
        ws.cell(row=r, column=2, value=s.name)
        ws.cell(row=r, column=3, value=getattr(s.owner, "phone", str(s.owner_id)))
        ws.cell(row=r, column=4, value=s.address)
        ws.cell(row=r, column=5, value=s.executor_kind)
        ws.cell(row=r, column=6, value=s.subscription_plan)
        ws.cell(row=r, column=7, value=s.subscription_paid_until.isoformat() if s.subscription_paid_until else "")
        ws.cell(row=r, column=8, value="Да" if s.is_active else "Нет")
        ws.cell(row=r, column=9, value="Да" if s.is_verified else "Нет")
        r += 1
    for col, w in {"A": 8, "B": 34, "C": 18, "D": 44, "E": 12, "F": 12, "G": 14, "H": 10, "I": 10}.items():
        ws.column_dimensions[col].width = w
    for rr in range(4, r):
        ws[f"A{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="erp_stations.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def station_detail(request: HttpRequest, station_id: int) -> HttpResponse:
    from apps.audit.models import AuditLog
    from apps.billing.models import PaymentIntent
    from apps.billing.models import Subscription
    from apps.billing.models import PaymentIntentStatus
    from apps.bookings.models import Booking
    from apps.reviews.models import Review
    from apps.stations.models import ServiceStation
    from django.db.models import Count, Sum
    from django.db.models.functions import TruncDate

    st = (
        ServiceStation.objects.select_related("owner", "district")
        .prefetch_related("categories", "car_brands")
        .get(pk=station_id)
    )
    sub = Subscription.objects.select_related("station").filter(station=st).first()

    # --- Payments filters (station-level) ---
    today = timezone.localdate()
    default_start = today - timedelta(days=30)
    raw_start = (request.GET.get("p_start") or "").strip()
    raw_end = (request.GET.get("p_end") or "").strip()
    raw_status = (request.GET.get("p_status") or "").strip()

    def _parse(raw: str):
        try:
            from datetime import date as _date

            return _date.fromisoformat(raw)
        except Exception:
            return None

    p_start = _parse(raw_start) or default_start
    p_end = _parse(raw_end) or today
    if p_end < p_start:
        p_start, p_end = p_end, p_start

    payments_qs = PaymentIntent.objects.select_related("subscription").filter(subscription__station=st)
    payments_period = payments_qs.filter(created_at__date__gte=p_start, created_at__date__lte=p_end)
    if raw_status:
        if raw_status in {s for s, _ in PaymentIntentStatus.choices}:
            payments_period = payments_period.filter(status=raw_status)

    payments = payments_period.order_by("-created_at")[:200]

    # Aggregates for the period (independent of raw_status)
    succeeded = payments_qs.filter(
        created_at__date__gte=p_start,
        created_at__date__lte=p_end,
        status=PaymentIntentStatus.SUCCEEDED,
    )
    failed_like = payments_qs.filter(
        created_at__date__gte=p_start,
        created_at__date__lte=p_end,
        status__in=[PaymentIntentStatus.FAILED, PaymentIntentStatus.CANCELED],
    )
    sum_succeeded = succeeded.aggregate(v=Sum("amount")).get("v") or 0
    sum_failed = failed_like.aggregate(v=Sum("amount")).get("v") or 0
    cnt_succeeded = succeeded.aggregate(c=Count("id")).get("c") or 0
    cnt_failed = failed_like.aggregate(c=Count("id")).get("c") or 0

    day_list = [p_end - timedelta(days=i) for i in range((p_end - p_start).days, -1, -1)]
    sums_by_day = dict(
        succeeded.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(v=Sum("amount"))
        .values_list("day", "v")
    )
    failed_cnt_by_day = dict(
        failed_like.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(c=Count("id"))
        .values_list("day", "c")
    )
    bookings = Booking.objects.select_related("client", "slot").filter(station=st).order_by("-created_at")[:200]
    reviews = (
        Review.objects.select_related("booking", "booking__client")
        .filter(station=st)
        .order_by("-created_at")[:200]
    )
    audit = AuditLog.objects.filter(object_type="stations.ServiceStation", object_id=st.id).order_by("-created_at")[:200]
    return render(
        request,
        "erp/station_detail.html",
        {
            "st": st,
            "sub": sub,
            "payments": payments,
            "payment_filters": {
                "start": p_start,
                "end": p_end,
                "status": raw_status,
            },
            "payment_status_choices": list(PaymentIntentStatus.choices),
            "payment_kpis": {
                "sum_succeeded": sum_succeeded,
                "sum_failed": sum_failed,
                "cnt_succeeded": cnt_succeeded,
                "cnt_failed": cnt_failed,
            },
            "payment_chart": {
                "labels_json": json.dumps([d.isoformat() for d in day_list]),
                "succeeded_amounts_json": json.dumps([float(sums_by_day.get(d, 0) or 0) for d in day_list]),
                "failed_counts_json": json.dumps([int(failed_cnt_by_day.get(d, 0) or 0) for d in day_list]),
            },
            "bookings": bookings,
            "reviews": reviews,
            "audit": audit,
            "subscription_plan_choices": [
                ("free", "Free"),
                ("basic", "Basic"),
            ],
        },
    )


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def station_activate(request: HttpRequest, station_id: int) -> HttpResponse:
    from apps.stations.models import ServiceStation

    st = ServiceStation.objects.get(pk=station_id)
    st.is_active = True
    st.save(update_fields=["is_active"])
    audit_log(
        request=request,
        event_type="erp.station",
        action="activate",
        obj=st,
        object_label=f"ServiceStation id={st.id}",
        payload={},
    )
    messages.success(request, "СТО активирована.")
    return redirect("erp:station_detail", station_id=st.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def station_deactivate(request: HttpRequest, station_id: int) -> HttpResponse:
    from apps.stations.models import ServiceStation

    st = ServiceStation.objects.get(pk=station_id)
    st.is_active = False
    st.save(update_fields=["is_active"])
    audit_log(
        request=request,
        event_type="erp.station",
        action="deactivate",
        obj=st,
        object_label=f"ServiceStation id={st.id}",
        payload={},
    )
    messages.success(request, "СТО выключена.")
    return redirect("erp:station_detail", station_id=st.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def station_set_subscription(request: HttpRequest, station_id: int) -> HttpResponse:
    from apps.stations.constants import SUBSCRIPTION_PLAN_BASIC, SUBSCRIPTION_PLAN_CHOICES, SUBSCRIPTION_PLAN_FREE
    from apps.stations.models import ServiceStation

    st = ServiceStation.objects.get(pk=station_id)
    old_plan = st.subscription_plan
    old_paid = st.subscription_paid_until.isoformat() if st.subscription_paid_until else None

    plan = (request.POST.get("plan") or "").strip()
    allowed = {SUBSCRIPTION_PLAN_FREE, SUBSCRIPTION_PLAN_BASIC}
    if plan not in allowed:
        messages.error(request, "Неверный тариф.")
        return redirect("erp:station_detail", station_id=st.id)

    paid_until_raw = (request.POST.get("paid_until") or "").strip()
    paid_until = None
    if paid_until_raw:
        try:
            from datetime import date as _date

            paid_until = _date.fromisoformat(paid_until_raw)
        except Exception:
            messages.error(request, "Неверная дата 'оплачено до'. Используй YYYY-MM-DD.")
            return redirect("erp:station_detail", station_id=st.id)

    st.subscription_plan = plan
    st.subscription_paid_until = paid_until
    st.save(update_fields=["subscription_plan", "subscription_paid_until"])

    audit_log(
        request=request,
        event_type="erp.station_subscription",
        action="set",
        obj=st,
        object_label=f"ServiceStation id={st.id}",
        payload={
            "from_plan": old_plan,
            "to_plan": plan,
            "from_paid_until": old_paid,
            "to_paid_until": paid_until.isoformat() if paid_until else None,
        },
    )
    messages.success(request, "Подписка обновлена.")
    return redirect("erp:station_detail", station_id=st.id)


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def station_bookings_xlsx(request: HttpRequest, station_id: int) -> HttpResponse:
    from apps.bookings.models import Booking
    from apps.stations.models import ServiceStation
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    st = ServiceStation.objects.get(pk=station_id)
    qs = Booking.objects.select_related("client", "slot").filter(station=st).order_by("-created_at")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    ws["A1"] = "МаБибип — ERP экспорт: Записи СТО"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"СТО: {st.id} {st.name}"
    headers = ["ID", "Клиент", "Слот", "Статус", "Телефон", "Создано"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for b in qs[:10000]:
        slot = ""
        if b.slot_id:
            slot = f"{b.slot.date} {b.slot.start_time}-{b.slot.end_time}"
        ws.cell(row=r, column=1, value=int(b.id))
        ws.cell(row=r, column=2, value=getattr(b.client, "phone", str(b.client_id)))
        ws.cell(row=r, column=3, value=slot)
        ws.cell(row=r, column=4, value=b.status)
        ws.cell(row=r, column=5, value=b.contact_phone)
        ws.cell(row=r, column=6, value=b.created_at.isoformat(sep=" ", timespec="minutes"))
        r += 1
    for col, w in {"A": 8, "B": 18, "C": 24, "D": 14, "E": 18, "F": 20}.items():
        ws.column_dimensions[col].width = w
    for rr in range(4, r):
        ws[f"A{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="erp_station_{st.id}_bookings.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def station_reviews_xlsx(request: HttpRequest, station_id: int) -> HttpResponse:
    from apps.reviews.models import Review, ReviewComplaint
    from apps.stations.models import ServiceStation
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    st = ServiceStation.objects.get(pk=station_id)
    reviews = (
        Review.objects.select_related("booking", "booking__client")
        .filter(station=st)
        .order_by("-created_at")[:10000]
    )
    complaints = (
        ReviewComplaint.objects.select_related("review")
        .filter(station=st)
        .order_by("-created_at")[:10000]
    )

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Reviews"
    ws1["A1"] = "МаБибип — ERP экспорт: Отзывы по СТО"
    ws1["A1"].font = Font(size=14, bold=True)
    ws1["A2"] = f"СТО: {st.id} {st.name}"
    headers = ["ID", "Клиент", "Рейтинг", "Статус", "Дата", "Текст", "Причина модерации"]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for rev in reviews:
        ws1.cell(row=r, column=1, value=int(rev.id))
        ws1.cell(row=r, column=2, value=rev.booking.client.phone if rev.booking_id else "")
        ws1.cell(row=r, column=3, value=int(rev.rating))
        ws1.cell(row=r, column=4, value=rev.moderation_status)
        ws1.cell(row=r, column=5, value=rev.created_at.date().isoformat())
        ws1.cell(row=r, column=6, value=(rev.text or "")[:2000])
        ws1.cell(row=r, column=7, value=rev.moderation_reason or "")
        r += 1
    for col, w in {"A": 8, "B": 18, "C": 8, "D": 16, "E": 12, "F": 80, "G": 40}.items():
        ws1.column_dimensions[col].width = w
    for rr in range(4, r):
        ws1[f"A{rr}"].alignment = Alignment(horizontal="right")
        ws1[f"C{rr}"].alignment = Alignment(horizontal="right")

    ws2 = wb.create_sheet(title="Complaints")
    ws2["A1"] = "МаБибип — ERP экспорт: Жалобы от СТО"
    ws2["A1"].font = Font(size=14, bold=True)
    ws2["A2"] = f"СТО: {st.id} {st.name}"
    headers2 = ["ID", "Review ID", "Причина", "Статус", "Создано", "Решено"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for cpl in complaints:
        ws2.cell(row=r, column=1, value=int(cpl.id))
        ws2.cell(row=r, column=2, value=int(cpl.review_id))
        ws2.cell(row=r, column=3, value=cpl.reason)
        ws2.cell(row=r, column=4, value=cpl.status)
        ws2.cell(row=r, column=5, value=cpl.created_at.isoformat(sep=" ", timespec="minutes"))
        ws2.cell(
            row=r,
            column=6,
            value=cpl.resolved_at.isoformat(sep=" ", timespec="minutes") if cpl.resolved_at else "",
        )
        r += 1
    for col, w in {"A": 8, "B": 10, "C": 60, "D": 14, "E": 20, "F": 20}.items():
        ws2.column_dimensions[col].width = w
    for rr in range(4, r):
        ws2[f"A{rr}"].alignment = Alignment(horizontal="right")
        ws2[f"B{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="erp_station_{st.id}_reviews_complaints.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def station_payments_xlsx(request: HttpRequest, station_id: int) -> HttpResponse:
    from apps.billing.models import PaymentIntent
    from apps.stations.models import ServiceStation
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    st = ServiceStation.objects.get(pk=station_id)
    intents = (
        PaymentIntent.objects.select_related("subscription", "subscription__station")
        .filter(subscription__station=st)
        .order_by("-created_at")[:20000]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws["A1"] = "МаБибип — ERP экспорт: Платежи по СТО"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"СТО: {st.id} {st.name}"
    headers = ["Дата", "Провайдер", "Статус", "Сумма", "Валюта", "Provider payment id", "Idempotency key"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
    r = 5
    for it in intents:
        ws.cell(row=r, column=1, value=it.created_at.isoformat(sep=" ", timespec="minutes"))
        ws.cell(row=r, column=2, value=it.provider)
        ws.cell(row=r, column=3, value=it.status)
        ws.cell(row=r, column=4, value=float(it.amount))
        ws.cell(row=r, column=5, value=it.currency)
        ws.cell(row=r, column=6, value=it.provider_payment_id)
        ws.cell(row=r, column=7, value=it.idempotency_key)
        r += 1
    for col, w in {"A": 20, "B": 12, "C": 14, "D": 12, "E": 10, "F": 24, "G": 22}.items():
        ws.column_dimensions[col].width = w
    for rr in range(4, r):
        ws[f"D{rr}"].alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="erp_station_{st.id}_payments.xlsx"'
    return resp


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def review_detail(request: HttpRequest, review_id: int) -> HttpResponse:
    from apps.reviews.models import Review, ReviewComplaint

    r = (
        Review.objects.select_related("booking", "booking__station", "booking__client")
        .get(pk=review_id)
    )
    complaints = list(
        ReviewComplaint.objects.select_related("station")
        .filter(review=r)
        .order_by("-created_at")[:200]
    )
    return render(request, "erp/review_detail.html", {"r": r, "complaints": complaints})


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def review_hide(request: HttpRequest, review_id: int) -> HttpResponse:
    from apps.reviews.models import ModerationStatus, Review

    r = Review.objects.select_related("booking", "booking__station").get(pk=review_id)
    r.moderation_status = ModerationStatus.HIDDEN
    r.moderation_reason = (request.POST.get("reason") or "Скрыто администратором").strip()[:300]
    r.save(update_fields=["moderation_status", "moderation_reason"])

    audit_log(
        request=request,
        event_type="erp.review_moderation",
        action="hide",
        obj=r,
        object_label=f"Review id={r.id} station={r.booking.station_id}",
        payload={"reason": r.moderation_reason},
    )
    return redirect("erp:review_detail", review_id=r.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def review_publish(request: HttpRequest, review_id: int) -> HttpResponse:
    from apps.reviews.models import ModerationStatus, Review

    r = Review.objects.select_related("booking", "booking__station").get(pk=review_id)
    r.moderation_status = ModerationStatus.OK
    r.moderation_reason = ""
    r.save(update_fields=["moderation_status", "moderation_reason"])

    audit_log(
        request=request,
        event_type="erp.review_moderation",
        action="publish",
        obj=r,
        object_label=f"Review id={r.id} station={r.booking.station_id}",
        payload={},
    )
    return redirect("erp:review_detail", review_id=r.id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def complaint_accept(request: HttpRequest, complaint_id: int) -> HttpResponse:
    from apps.reviews.models import ComplaintStatus, ModerationStatus, ReviewComplaint

    c = ReviewComplaint.objects.select_related("review", "review__booking", "review__booking__station").get(
        pk=complaint_id
    )
    # accept complaint = hide review + resolve complaint
    c.review.moderation_status = ModerationStatus.HIDDEN
    if not (c.review.moderation_reason or "").strip():
        c.review.moderation_reason = "Скрыто по жалобе СТО"
    c.review.save(update_fields=["moderation_status", "moderation_reason"])

    c.status = ComplaintStatus.RESOLVED
    c.resolved_at = timezone.now()
    c.save(update_fields=["status", "resolved_at"])

    audit_log(
        request=request,
        event_type="erp.review_complaint",
        action="accept",
        obj=c,
        object_label=f"Complaint id={c.id} review={c.review_id} station={c.station_id}",
        payload={"review_hidden": True},
    )
    return redirect("erp:review_detail", review_id=c.review_id)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def complaint_reject(request: HttpRequest, complaint_id: int) -> HttpResponse:
    from apps.reviews.models import ComplaintStatus, ReviewComplaint

    c = ReviewComplaint.objects.select_related("review", "review__booking", "review__booking__station").get(
        pk=complaint_id
    )
    # reject complaint = resolve complaint without hiding review
    c.status = ComplaintStatus.RESOLVED
    c.resolved_at = timezone.now()
    c.save(update_fields=["status", "resolved_at"])

    audit_log(
        request=request,
        event_type="erp.review_complaint",
        action="reject",
        obj=c,
        object_label=f"Complaint id={c.id} review={c.review_id} station={c.station_id}",
        payload={"review_hidden": False},
    )
    return redirect("erp:review_detail", review_id=c.review_id)


def _support_ticket_detail_redirect(ticket_id: int, next_q: str) -> HttpResponse:
    from django.urls import reverse

    url = reverse("erp:support_ticket_detail", args=[ticket_id])
    if next_q.strip():
        return redirect(f"{url}?{next_q}")
    return redirect(url)


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def support_list(request: HttpRequest) -> HttpResponse:
    from apps.support.models import SupportTicket, SupportTicketStatus
    from apps.support.unread import support_unread_tickets_for_staff_qs

    status = (request.GET.get("status") or "").strip()
    start_d, end_d = _parse_range(request)
    unread_only = (request.GET.get("unread") or "").strip() in ("1", "true", "yes")

    qs = SupportTicket.objects.select_related("user").order_by("-updated_at", "-pk")
    if status:
        qs = qs.filter(status=status)
    qs = qs.filter(created_at__date__gte=start_d, created_at__date__lte=end_d)
    if unread_only:
        qs = qs.filter(pk__in=support_unread_tickets_for_staff_qs().values_list("pk", flat=True))

    tickets = list(qs[:500])
    staff_unread_ids = set(support_unread_tickets_for_staff_qs().values_list("pk", flat=True))
    return render(
        request,
        "erp/support_list.html",
        {
            "tickets": tickets,
            "status": status,
            "start": start_d.isoformat(),
            "end": end_d.isoformat(),
            "status_choices": list(SupportTicketStatus.choices),
            "unread_only": unread_only,
            "staff_unread_ticket_ids": staff_unread_ids,
        },
    )


@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def support_ticket_detail(request: HttpRequest, ticket_id: int) -> HttpResponse:
    from apps.support.models import SupportMessage, SupportTicket
    from apps.support.unread import mark_ticket_read_by_staff

    ticket = get_object_or_404(
        SupportTicket.objects.select_related("user"),
        pk=ticket_id,
    )
    mark_ticket_read_by_staff(ticket.pk)
    messages_list = list(
        SupportMessage.objects.select_related("author")
        .filter(ticket=ticket)
        .order_by("created_at", "pk")
    )
    back_query = request.GET.urlencode()
    return render(
        request,
        "erp/support_ticket_detail.html",
        {"ticket": ticket, "messages_list": messages_list, "back_query": back_query},
    )


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def support_ticket_reply(request: HttpRequest, ticket_id: int) -> HttpResponse:
    from apps.support.models import SupportMessage, SupportTicket, SupportTicketStatus

    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    body = (request.POST.get("body") or "").strip()
    next_q = (request.POST.get("next_q") or "").strip()
    if not body:
        messages.error(request, "Введите текст ответа.")
        return _support_ticket_detail_redirect(ticket_id, next_q)

    SupportMessage.objects.create(
        ticket=ticket,
        author=request.user,
        body=body,
        is_staff_reply=True,
        is_system_auto=False,
    )
    if ticket.status == SupportTicketStatus.OPEN:
        ticket.status = SupportTicketStatus.IN_PROGRESS
        ticket.save(update_fields=["status"])

    audit_log(
        request=request,
        event_type="erp.support_reply",
        action="create",
        obj=ticket,
        object_label=f"SupportTicket id={ticket.id}",
        payload={"message_len": len(body)},
    )
    messages.success(request, "Ответ сохранён.")
    return _support_ticket_detail_redirect(ticket_id, next_q)


@require_POST
@user_passes_test(_is_erp_admin, login_url="/accounts/login/")
def support_ticket_set_status(request: HttpRequest, ticket_id: int) -> HttpResponse:
    from apps.support.models import SupportTicket, SupportTicketStatus

    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    raw = (request.POST.get("status") or "").strip()
    next_q = (request.POST.get("next_q") or "").strip()
    allowed = {
        SupportTicketStatus.RESOLVED,
        SupportTicketStatus.CLOSED,
        SupportTicketStatus.OPEN,
    }
    if raw not in allowed:
        messages.error(request, "Недопустимый статус.")
        return _support_ticket_detail_redirect(ticket_id, next_q)

    old = ticket.status
    ticket.status = raw
    ticket.save(update_fields=["status"])

    audit_log(
        request=request,
        event_type="erp.support_ticket_status",
        action="set",
        obj=ticket,
        object_label=f"SupportTicket id={ticket.id}",
        payload={"from": old, "to": raw},
    )
    messages.success(request, "Статус обновлён.")
    return _support_ticket_detail_redirect(ticket_id, next_q)

